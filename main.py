import openpyxl
import pandas as pd
from pyautogui import size
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import subprocess
import shutil
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from bs4 import BeautifulSoup
import time
import datetime
import pyautogui
import pyperclip
import csv
import sys
import os
import math
import requests
import re
import random
import chromedriver_autoinstaller
from PyQt5.QtWidgets import QWidget, QApplication,QTreeView,QFileSystemModel,QVBoxLayout,QPushButton,QInputDialog,QLineEdit,QMainWindow,QMessageBox,QFileDialog,QTextEdit
from PyQt5.QtCore import QCoreApplication
from selenium.webdriver import ActionChains
from datetime import datetime,date,timedelta
import numpy
import datetime
from window import Ui_MainWindow
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
import pybase64
import bcrypt
import http.client
import json
import pprint
import firebase_admin
from firebase_admin import credentials
from firebase_admin import db


def get_token(price, productNo, api_id,api_pw):
    time_now = datetime.datetime.now() - datetime.timedelta(seconds=3)
    time_now_stamp = math.ceil(datetime.datetime.timestamp(time_now) * 1000)
    # print(time_now)
    # print(time_now_stamp)

    clientId = api_id  # client id
    clientSecret = api_pw  # client pw
    # clientId=self.clientid
    # clientSecret=self.clientkey
    # timestamp = 1643961623299
    timestamp = time_now_stamp
    # 밑줄로 연결하여 password 생성
    password = clientId + "_" + str(timestamp)
    # bcrypt 해싱
    hashed = bcrypt.hashpw(password.encode('utf-8'), clientSecret.encode('utf-8'))
    # base64 인코딩
    result = pybase64.standard_b64encode(hashed).decode('utf-8')
    # print(result)
    params = {
        "client_id": clientId,
        "timestamp": time_now_stamp,
        "client_secret_sign": result,
        "grant_type": "client_credentials",
        "type": "SELF"
    }
    res = requests.post('https://api.commerce.naver.com/external/v1/oauth2/token', params=params)
    res.raise_for_status()

    token = eval(res.text)['access_token']
    conn = http.client.HTTPSConnection("api.commerce.naver.com")
    headers = {'Authorization': "Bearer {}".format(token)}
    conn.request("GET", "/external/v2/products/channel-products/{}".format(productNo), headers=headers)
    res = conn.getresponse()
    data = res.read()

    result = data.decode("utf-8")


    json_new_result = json.loads(result)
    # pprint.pprint(json_new_result)
    if json_new_result['originProduct']['deliveryInfo']['deliveryCompany']=="":
        json_new_result['originProduct']['deliveryInfo']['deliveryCompany']="HANJIN"
    # print("택배사는:",json_new_result['originProduct']['deliveryInfo']['deliveryCompany'])
    origin_price = int(json_new_result['originProduct']['salePrice'])
    json_new_result['originProduct']['salePrice']=price
    json_new_result['originProduct']['detailContent'] = ""


    file_path = 'result.json'
    with open(file_path, 'w') as f:
        json.dump(json_new_result, f)

    token_path = 'token.txt'
    f = open(token_path, 'w')
    f.write(token)
    f.close()
    print("겟토큰완료")
def change_price(productNo):
    token_path = 'token.txt'
    with open(token_path) as f:
        lines = f.readlines()
        token = lines[0].strip()

    file_path = 'result.json'
    with open(file_path, 'r') as f:
        data = json.load(f)

    headers = {
        'Authorization': token,
        'content-type': "application/json"
    }

    # pprint.pprint(data)
    # print("PUT요청 보내기")
    res = requests.put(
        'https://api.commerce.naver.com/external/v2/products/channel-products/{}'.format(productNo),
        data=json.dumps(data), headers=headers)
    # print("PUT요청 완료")
    # res.raise_for_status()
    result = res.status_code
    print('가격변경코드전송상태:', result)
def find_price(productNo):
    token_path = 'token.txt'
    with open(token_path) as f:
        lines = f.readlines()
        token = lines[0].strip()
    # print(token)

    file_path = 'result.json'
    with open(file_path, 'r') as f:
        data = json.load(f)
    # print(data)

    headers = {'Authorization': "Bearer {}".format(token)}

    res = requests.get(
        'https://api.commerce.naver.com/external/v2/products/channel-products/{}'.format(productNo),
        headers=headers)
    res.raise_for_status()
    # pprint.pprint(json.loads(res.text))
    res_dic = json.loads(res.text)
    name = res_dic['originProduct']['name']

    try:
        discount_price = int(
            res_dic['originProduct']['customerBenefit']['immediateDiscountPolicy']['mobileDiscountMethod']['value'])
        price = int(json.loads(res.text)['originProduct']['salePrice'] - discount_price)
    except:
        price = int(json.loads(res.text)['originProduct']['salePrice'])
    # print("이름은:", name)
    return name, price
def load_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_list=wb.get_sheet_names()
    ws = wb[sheet_list[1]]
    no_row = ws.max_row
    print("행갯수:", no_row)
    info_list = []
    for i in range(2, no_row + 1):
        # print(i,"번째 행 정보 가져오는 중...")
        productNo = ws.cell(row=i, column=1).value
        if productNo==None:
            productNo=""
        name = ws.cell(row=i, column=2).value
        if name=="" or name==None:
            break
        url_catalog = ws.cell(row=i, column=3).value
        if url_catalog==None:
            url_catalog=""
        if url_catalog:
            url_catalog=url_catalog.replace("https://search",'https://msearch')
        url_target=ws.cell(row=i, column=4).value
        if url_target==None:
            url_target=""
        price_low = ws.cell(row=i, column=5).value
        price_tic = int(ws.cell(row=i, column=6).value)
        switch = ws.cell(row=i, column=7).value

        info = [productNo, name,url_catalog, url_target,price_low,price_tic,switch]

        info_list.append(info)
    print("info_list:",info_list)
    return info_list
def get_catalog_price(url, store_name,exception_list):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/108.0.0.0 Safari/537.36"}


    res = requests.get(url, headers=headers)
    res.raise_for_status()
    position_fr = res.text.find("{")
    position_rr = res.text.rfind("}")
    result_raw = res.text[position_fr:position_rr + 1]
    result = json.loads(result_raw)
    result_list = result['props']['pageProps']['dehydratedState']['queries']

    mall_total_list = []
    # mall_useless=['11번가','G마켓','옥션','쿠팡','위메프','롯데','템스윈공식몰','인터파크','인터파크쇼핑']
    mall_useless=[]
    if len(exception_list)>=1:
        mall_useless.extend(exception_list)
        print("오픈몰+제외몰:",mall_useless)

    for index, result_elem in enumerate(result_list):
        try:
            mall_list = result_elem['state']['data']['pages'][0]['products']
        except:
            # print("없음")
            mall_list = []
        for mall_elem in mall_list:
            if mall_elem['mallName'] in mall_useless:
                continue
            # print("몰이름:", mall_elem['mallName'], "가격:", mall_elem['mobilePrice'])
            data = [mall_elem['mallName'], int(mall_elem['mobilePrice'])]
            mall_total_list.append(data)
    print("mall_total_list:", mall_total_list)

    first_flag = True
    for mall_total_elem in mall_total_list:
        price_mall = mall_total_elem[1]
        name_mall = mall_total_elem[0]
        # print("몰가격:", price_mall, "몰이름:", name_mall)
        if first_flag == True:
            least_price = price_mall
            if name_mall.find(store_name) >= 0:
                is_first = True
                print("1등여부:", is_first)
            else:
                is_first = False
            first_flag = False
        elif first_flag == False:
            second_price = price_mall
            break

    return least_price, second_price, is_first
def get_target_price(url):
    url = 'https://smartstore.naver.com/1cc/products/7190863120?NaPm=ct%3Dlfm3pj5k%7Cci%3D743a40b6df75b561265ff23978ea1f990e632c4a%7Ctr%3Dslsc%7Csn%3D4367970%7Chk%3D3b9234ab4ccb9ace4a557ccedc0848348b46b343'
    headers = {
        'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 9_3_2 like Mac OS X) AppleWebKit/601.1.46 (KHTML, like Gecko) Version/9.0 Mobile/13F69 Safari/601.1',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Charset': 'ISO-8859-1,utf-8;q=0.7,*;q=0.3',
        'Accept-Encoding': 'none',
        'Accept-Language': 'en-US,en;q=0.8',
        'Connection': 'keep-alive'}
    res = requests.get(url, headers=headers)
    soup = BeautifulSoup(res.text, 'lxml')
    head = soup.find('head')
    script = head.find_all('script')[0]
    position_fr = str(script).find("{")
    position_rr = str(script).rfind("}")
    result_raw = str(script)[position_fr:position_rr + 1]
    result = int(json.loads(result_raw)['offers']['price'])
    print("타겟가격:",result)
    return result
def load_store(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_list = wb.get_sheet_names()
    ws = wb[sheet_list[1]]
    # ws = wb.active
    no_row = ws.max_row
    print("행갯수:", no_row)
    info_list = []
    for i in range(2, no_row + 1):
        # print(i, "번째 행 정보 가져오는 중...")
        storeName = ws.cell(row=i, column=13).value
        if storeName == "" or storeName == None:
            break
        info = storeName
        info_list.append(info)
    print("exception_list:", info_list)
    return info_list

def get_key(first_flag):
    print("111")
    if first_flag==True:
        secret={
          "type": "service_account",
          "project_id": "hanbaik",
          "private_key_id": "14ac8e67260883ee2c095ab58e32b27c41df23dd",
          "private_key": "-----BEGIN PRIVATE KEY-----\nMIIEvgIBADANBgkqhkiG9w0BAQEFAASCBKgwggSkAgEAAoIBAQDbX5A3/CzD78yf\n68NUAGB8fGyDAwY8wk2ngS2aYHNCNsnk2MVJQtCZTEo9PaRGG+xrlXKpJBMFQdmZ\nIctMpff9E+q0PDDASSgIjhbmcRT4vfMe2gRTeMISdezirXkbzGFKAKvjhQdtF+y2\n3y/KIpHDJeyesTKaAdOCAcFj2G/V3JucXNhxrrwgFhHHqBGET0c+BDTh/9DkX2M7\nav4hSmkdGGma2qsFdc1LzfWEwrN6Izqv4VGAFxSbhTL6/eyAxOc9jqbho3jVn9Ix\nXSpLsI64LwImJWPBCigy0Hoo9PMxyBtocKygDaK9Ii79KCF1cvIUaDZY1nu6Wns9\nknN21vR/AgMBAAECggEAOEFLVHAAcsZ9rMzirBnkpEer8/TElrQlAb1omlv9co9m\nOp80CFNv9r/PkeDKzYe/mt8aJjGwBvsZ9+Dop2EwNN/0of+FaOnZsEfvq2x4OklL\nnS+/SECBVAaVlpNxqSVCFZ4SdifY/arS8xpMwQMYafjBsvgxx7iMKpyUoSwRkb+l\nQtMagdhTEvJl7QwLnCpLgGFWs/3GjG8M5RD8CdHjGqxTaigYOkjU6bD00csIHkbI\nGveP585Qnf1rSbjZ83blk2S6gRqb9+OW4J8Z5Okwm4Dmb3LXkHgGCoVNT6f1sFsZ\nhsm1L8h7gztJwVmpoNTSDr4xQKiXQRGwb4KRfoHuMQKBgQDv730N4itlFadYxZdw\nktskJhjKs4N8LRCO6dy7uoU1GNIuiy+xFXhcGNSSnSM6JarbRZaVU0H4qtd9KtZ0\nl0fIO6jZLTQKyDJJEZifVYEKFXW811yfYBtvObh00gt40/26dajqct2oulIQxzq6\n7NVpa9EEdEaCRq2kTa+c8SI60wKBgQDqD6MdUgV5S3vmn4KW4uxT60ZtucvWT/H7\nOJNzvjT3YdeMhKG1gbB/wOn+W5tNMj5zzOH7Zngu+fCATHNU/ndohAp0cas3itw8\nIpIlNRR45Vvdcjn19Ml0qVh87oS9O6B9njCXC56jKKOAc5kWjgA6VsHJFEYBh8SU\nYe7TdZk8JQKBgEE9gWdxBBOsW6CLua3mgKfHpB4ZybrOFh6GAHsbMHVLlnsJZaJl\nECEar1JeX+HDtD2DInrf9KRE7+sc5ss1B1OuxS6oV+pGnUW4/yL0AO5Y/3alqI29\neDg6HanGI1BrdCZrL87wBM2IPCBLy/BfzXeo1WC8rR9nUHfIl+O4vXH5AoGBANgr\nDvuyV+nZRBoP8XzHIXqzzTzjnpVVCmh5rPz1i1d6HqfhirPmjgq/MZzAICNgpvsu\nGvujfJXuMidb9BxoVAHMCRfYL0hBz/sd9pm0dy7crUZNC6jTpgc/q8DeTOu0GRpL\nMhceHSoVC0RD/vwss5stqxW5ypn5OR3NgNP9RUOdAoGBALglxCsn4/O3aDB4CotP\nbzdQWSLneFXEUz3iQsTIvdWMUnT7onwWVkECD1DmcpngFy+/b2wEHp4uOrKvb5Nx\nLbrzpppHTBOmtEnh076T5K3JtP+ZgdmxgkjX3csRoQBk1VhdK+l1Cx1j9krct9rB\nLpDPBYw2Uyt/OucarZjmMf56\n-----END PRIVATE KEY-----\n",
          "client_email": "firebase-adminsdk-nkdhd@hanbaik.iam.gserviceaccount.com",
          "client_id": "103946102534201211018",
          "auth_uri": "https://accounts.google.com/o/oauth2/auth",
          "token_uri": "https://oauth2.googleapis.com/token",
          "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
          "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/firebase-adminsdk-nkdhd%40hanbaik.iam.gserviceaccount.com"
        }

        cred = credentials.Certificate(secret)
        firebase_admin.initialize_app(cred,{
            'databaseURL': "https://hanbaik-default-rtdb.firebaseio.com/"
        })
    print(db)

    password=db.reference().get()['users']['password']
    print(password)
    id = db.reference().get()['users']['id']
    print("id:",id,'password:',password)
    # ref=db.reference().get()
    # result_password=ref['users']

    return id,password


class Thread(QThread):
    # 초기화 메서드 구현
    def __init__(self, parent,file_path,store_name,api_id,api_pw,start_row,end_row):  # parent는 WndowClass에서 전달하는 self이다.(WidnowClass의 인스턴스)
        super().__init__(parent)
        self.parent = parent  # self.parent를 사용하여 WindowClass 위젯을 제어할 수 있다.
        self.running_flag=True
        self.file_path=file_path
        self.store_name=store_name
        self.api_id=api_id
        self.api_pw=api_pw
        self.start_row=start_row
        self.end_row=end_row



    def run(self):

        while True:
            print("2222")
            if self.running_flag==False or len(self.file_path)==0:
                break


            # -----------------실행부위-------------
            while True:
                try:
                    clock_now = datetime.datetime.now()
                    clock_now_hour = clock_now.strftime("%H")
                    clock_now_minute = clock_now.strftime("%M")
                    clock_now_second = clock_now.strftime("%S")
                    print("현재시간: {}시 {}분 {}초".format(clock_now_hour, clock_now_minute, clock_now_second))
                    info_list = load_excel(self.file_path)
                    print("엑셀읽어오기완료")
                    if int(clock_now_minute):
                        print("조정시작")

                        # self.diff=int(1/len(info_list)*100)
                        # self.num=0
                        exception_list=load_store(self.file_path)


                        for index, info in enumerate(info_list):
                            if index<self.start_row or index>self.end_row:
                                continue

                            if self.running_flag==False:
                                break
                            productNo = info[0]
                            if productNo=="":
                                continue
                            name = info[1]
                            url = info[2]
                            if url=="":
                                continue
                            url_target=info[3]
                            price_low = info[4]
                            price_tic = info[5]
                            switch = info[6]

                            if switch==0:
                                print("조정안하는 상품 스킵함")
                                continue

                            if url == "" or url == None:
                                print("url없어서 넘어감")
                                continue

                            nownow = datetime.datetime.now()
                            nownow = nownow.strftime("%Y%m%d_%H%M")
                            text="{}번째 상품 크롤링 중.. 번호 : {} / {}".format(index+2,productNo,name)

                            print(text)
                            try:
                                least_price, second_price, is_first = get_catalog_price(url,self.store_name,exception_list)
                            except:
                                print("카탈로그 조회 에러로 건너뜀")
                                print("=================================")
                                time.sleep(0.5)
                                continue
                            if productNo == "" or productNo == None:
                                print("상품번호없어서 넘어감")
                                continue
                            try:
                                get_token(price_low,productNo,self.api_id,self.api_pw)
                            except:
                                print("토큰 발급 에러로 건너뜀")
                                print("=================================")
                                time.sleep(0.5)
                                continue
                            # print("현재가격찾기")
                            name, current_price = find_price(productNo)
                            # print("최저가격은:", least_price,"둘째가격은:",second_price, "현재가격은:", current_price, "1등여부:", is_first)
                            # print("카탈로그는:", url)

                            if switch == 0:
                                print("가격조절불가")
                            if switch >= 1:
                                if switch==2:
                                    print("타겟팅모드")
                                    least_price=get_target_price(url_target)
                                    second_price=least_price+10
                                    is_first=False
                                    print("최저가격은:", least_price, "둘째가격은:", second_price, "현재가격은:", current_price,"하한가는:",price_low,"1등여부:", is_first)
                                    print("카탈로그는:", url)
                                else:
                                    print("일반모드")
                                    print("최저가격은:", least_price, "둘째가격은:", second_price, "현재가격은:", current_price,"하한가는:",price_low,"1등여부:", is_first)
                                    print("카탈로그는:", url)
                                if is_first == True:
                                    if least_price==10 and price_low<=10:
                                        if second_price>current_price+10:
                                            print("2위 -10원으로 변경2")
                                            get_token(second_price - price_tic, productNo,self.api_id,self.api_pw)
                                            change_price(productNo)
                                        else:
                                            print("변경해당없음1")
                                    elif current_price>price_low:
                                        if current_price==second_price-price_tic:
                                            print("가격 기 지정 완료")
                                        elif current_price!=second_price-price_tic:
                                            print("2위 -10원으로 변경1")
                                            get_token(second_price-price_tic, productNo,self.api_id,self.api_pw)
                                            change_price(productNo)
                                            print("변경완료")
                                    elif current_price==price_low:
                                        if second_price>=current_price+price_tic:
                                            print("2위 -10원으로 변경2")
                                            get_token(second_price - price_tic, productNo,self.api_id,self.api_pw)
                                            change_price(productNo)
                                        else:
                                            print("하한가라 변경안함")
                                    else:
                                        print("하한가 보다 낮아서 상향함")
                                        get_token(price_low, productNo,self.api_id,self.api_pw)
                                        change_price(productNo)
                                        print("변경완료")

                                elif is_first == False:
                                    if least_price==10 and price_low<=10:
                                        if second_price>current_price+10:
                                            print("2위 틱으로 변경2")
                                            get_token(second_price - price_tic, productNo,self.api_id,self.api_pw)
                                            change_price(productNo)
                                        else:
                                            print("10원유지")
                                            get_token(10, productNo,self.api_id,self.api_pw)
                                            print("가격변경시도")
                                            change_price(productNo)
                                            print("변경완료")
                                    elif current_price>price_low:
                                        print("1위 뺏기")
                                        if least_price-price_tic<=price_low: # 1위탈환할 때 하한가 보다 낮으면 하한가로 셋팅
                                            print("하한가까지 밖에 못 내려감")
                                            price_change=price_low
                                        else:
                                            price_change = least_price - price_tic
                                        get_token(price_change, productNo, self.api_id, self.api_pw)
                                        # print("토큰 발행완료")
                                        change_price(productNo)
                                        print("변경완료")
                                    elif current_price==price_low:
                                        if second_price>current_price+price_tic:
                                            print("2위 -틱으로 변경2")
                                            get_token(second_price - price_tic, productNo,self.api_id,self.api_pw)
                                            change_price(productNo)
                                        else:
                                            print("하한가라 변경안함")
                                    else:
                                        print("하한가보다 낮아서 상향함")
                                        get_token(price_low, productNo,self.api_id,self.api_pw)
                                        change_price(productNo)
                                        print("변경완료")

                            time.sleep(0.5)
                            print("===================================")
                        if self.running_flag==False:
                            break
                    time.sleep(1)
                except:
                    print('토큰 에러로 한텀 쉬기')
                    time.sleep(5)

    def stop(self):
        self.running_flag=False
        self.quit()
        print("1111213123",self.running_flag)


class Example(QMainWindow,Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.path="C:"
        self.index=None
        self.setupUi(self)
        self.setSlot()
        self.show()
        self.file_path=""
        self.running_flag=True
        QApplication.processEvents()
        self.first_flag=True
        self.lineEdit.setText("hanbaik0422")
        self.lineEdit_6.setText("gksqor1004")
        self.login_flag=False
        self.lineEdit_7.setText("1")
        self.lineEdit_8.setText("999999")


    def start(self):
        self.start_row=int(self.lineEdit_7.text())-2
        self.end_row = int(self.lineEdit_8.text())-2

        if self.login_flag==True:
            self.api_id=self.lineEdit_4.text()
            self.api_pw=self.lineEdit_5.text()
            self.file_path = self.lineEdit_3.text()
            self.store_name=self.lineEdit_2.text()
            print("11")
            if len(self.file_path)==0:
                QMessageBox.information(self, "에러", "엑셀 파일을 Import 하세요")
            else:
                self.x=Thread(self, self.file_path,self.store_name,self.api_id,self.api_pw,self.start_row,self.end_row)
                self.x.start()
        else:
            QMessageBox.information(self, "에러", "로그인에 성공 후 실행하세요")
    def stop(self):
        self.running_flag=False
        # self.x = Thread(self, self.file_path,self.clientid,self.clientkey)
        # self.x.stop()
        self.x.terminate()
    def on_login(self):

        firebase_id,firebase_password=get_key(self.first_flag)

        self.first_flag = False
        insert_id = self.lineEdit.text()
        insert_pw = self.lineEdit_6.text()


        if firebase_id==insert_id and firebase_password==insert_pw:
            print("로그인에 성공하였습니다.")
            self.login_flag=True

        if self.login_flag==True:
            self.textEdit.append("로그인에 성공 하였습니다.")
        else:
            self.textEdit.append("로그인에 실패 하였습니다.")
    def setSlot(self):
        pass
    def setIndex(self,index):
        pass
    def quit(self):
        QCoreApplication.instance().quit()
    def search(self):
        fname = QFileDialog.getOpenFileName(self, "Open file", './')
        print(fname[0])
        self.file_path=fname[0]
        self.lineEdit_3.setText(fname[0])
        wb=openpyxl.load_workbook(fname[0])
        ws=wb.active
        self.id=ws.cell(row=2,column=9).value
        self.pw=ws.cell(row=2,column=10).value
        self.store_name = ws.cell(row=2, column=11).value
        self.lineEdit_4.setText(self.id)
        self.lineEdit_5.setText(self.pw)
        self.lineEdit_2.setText(self.store_name)
app=QApplication([])
ex=Example()
sys.exit(app.exec_())



if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())


